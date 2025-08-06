import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
from django.urls import reverse
from .models import Capacity,Free,CapacitySection,TypePoints,CourseYear
from AcademicInfoApp.models import Year,Section,Course,Degree,Language
from UsersApp.models import Professor
from .forms import UploadForm
from django.shortcuts import render, redirect,get_object_or_404
from django.contrib import messages
from django.db.models import Sum
import re

def sanitize_sheet_title(title):
    # Invalid characters for Excel sheet names
    invalid_chars = r'[\\/:*?\"<>|]'
    return re.sub(invalid_chars, '_', title)

## CAPACITYPROFESSORS EXCEL

def generate_capacityprofessor_excel(request,year_id):
    try:
        # Validate and fetch the selected year
        try:
            selected_year = Year.objects.get(pk=int(year_id))
        except (Year.DoesNotExist, ValueError, TypeError):
            messages.error(request, f"Selecciona un curs acadèmic que existeixi: {str(e)}")
            return redirect("capacityprofessor_list")


        # Fetch all sections and professors for the selected year
        all_sections = Section.objects.all()
        professor_data = []

        capacities = Capacity.objects.filter(Year_id=selected_year.idYear).order_by('Professor__family_name')

        for capacity in capacities:
            professor = capacity.Professor
            free_points = Free.objects.filter(Professor=professor, Year=selected_year).aggregate(free_points=Sum('PointsFree'))['free_points'] or 0
            section_points = [(section.LetterSection, 0) for section in all_sections]

            for section_entry in CapacitySection.objects.filter(Professor=professor, Year=selected_year):
                section_letter = section_entry.Section.LetterSection
                section_points = [(letter, section_entry.Points if letter == section_letter else points) for letter, points in section_points]

            section_points_sum = sum(points for _, points in section_points)
            balance = capacity.Points - free_points - section_points_sum

            professor_data.append({
                'id': professor.idProfessor,
                'name': professor.name,
                'family_name': professor.family_name,
                'total_points': capacity.Points,
                'free_points': free_points,
                'sections': dict(section_points),
                'balance': balance,
            })

        # Generate Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sanitize_sheet_title(f"Capacitat_docent_{selected_year.Year}")

        # Header row
        headers = ['ID del Professor', 'Nom', 'Cognom', 'Punts totals', "Punts d'alliberació"] + [section.LetterSection for section in all_sections] + ['Balanç']
        sheet.append(headers)

        # Data rows
        for data in professor_data:
            row = [
                data['id'],
                data['name'],
                data['family_name'],
                data['total_points'],
                data['free_points'],
            ]
            row.extend([data['sections'].get(section.LetterSection, 0) for section in all_sections])
            row.append(data['balance'])
            sheet.append(row)

        # Adjust column widths
        for col in sheet.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
            sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length

        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="capacitat_docent_{selected_year.Year}.xlsx"'
        workbook.save(response)

        return response
    except Exception as e:
        messages.error(request, f"Error al generar el fitxer Excel: {str(e)}")
        return redirect("capacityprofessor_list")

def upload_capacityprofessor_excel(request):
    return 

def generate_courseyear_excel(request,year_id):
    try:
        # Validate and fetch the selected year
        try:
            selected_year = Year.objects.get(pk=int(year_id))
        except (Year.DoesNotExist, ValueError, TypeError):
            messages.error(request, f"Selecciona un curs acadèmic que existeixi: {str(e)}")
            return redirect("courseyear_list")
        
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title =  sanitize_sheet_title(f"Encàrrec docent curs {selected_year.Year}")

        # Define headers
        headers = [
            ("Id encàrrec docent", 15),
            ("Curs acadèmic", 15),
            ("Titulació",30),
            ("Codi assignatura",15),
            ("Assignatura", 30),
            ("Semestre", 8),
            ("Punts A", 8),
            ("Punts B", 8),
            ("Punts C", 8),
            ("Punts D", 8),
            ("Punts E", 8),
            ("Punts F", 8),
            ("Suma",10),
            ("Idioma", 15),
            ("Comentari", 30)
        ]

        # Write headers to the first row
        for col_num, (header, width) in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        course_year_entries = CourseYear.objects.filter(Year=selected_year).order_by(
            "Course__Degree__NameDegree", "Course__NameCourse"
        )
        for row_num, course_year in enumerate(course_year_entries, start=2):
            sheet.cell(row=row_num, column=1, value=course_year.idCourseYear).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=2, value=course_year.Year.Year).alignment = Alignment(horizontal="left")  
            sheet.cell(row=row_num, column=3, value=course_year.Course.Degree.NameDegree).alignment = Alignment(horizontal="left")  
            sheet.cell(row=row_num, column=4, value=course_year.Course.CodeCourse).alignment = Alignment(horizontal="left")  
            sheet.cell(row=row_num, column=5, value=course_year.Course.NameCourse).alignment = Alignment(horizontal="left")  
            sheet.cell(row=row_num, column=6, value=course_year.Semester).alignment = Alignment(horizontal="left")  
            sheet.cell(row=row_num, column=7, value=course_year.PointsA).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=8, value=course_year.PointsB).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=9, value=course_year.PointsC).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=10, value=course_year.PointsD).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=11, value=course_year.PointsE).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=12, value=course_year.PointsF).alignment = Alignment(horizontal="left")
            
            # Sum of points with left alignment
            sheet.cell(row=row_num, column=13, value=sum([
                course_year.PointsA or 0,
                course_year.PointsB or 0,
                course_year.PointsC or 0,
                course_year.PointsD or 0,
                course_year.PointsE or 0,
                course_year.PointsF or 0
            ])).alignment = Alignment(horizontal="left")
            
            sheet.cell(row=row_num, column=14, value=course_year.Language.Language if course_year.Language else '').alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=15, value=course_year.Comment).alignment = Alignment(horizontal="left")
        
        # Generate a timestamped file name
        current_datetime = datetime.now().strftime("%d%m%Y_%H%M%S")
        filename = f"Encarrec docent {selected_year.Year}_{current_datetime}.xlsx"

        # Set up the HTTP response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        workbook.save(response)

        return response
    except Exception as e:
        messages.error(request, f"Error al generar el fitxer Excel: {str(e)}")
        return redirect("courseyear_list")

def upload_courseyear_excel(request):
    redirect_year = None
    if request.method == "POST":
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["file"]
            error_occurred = False
            
            try:
                # Open the Excel workbook
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                # Define the expected headers
                expected_columns = ["Id encàrrec docent","Curs acadèmic","Titulació","Codi assignatura","Assignatura","Semestre","Punts A",
                        "Punts B","Punts C","Punts D","Punts E","Punts F", "Suma","Idioma","Comentari",]
                
                header_row = [cell.value for cell in sheet[1]]

                # Check columns
                if not all(col in header_row for col in expected_columns):
                    messages.error(request, "El fitxer no té les columnes esperades. Comprova que el fitxer segueixi les dades com el excel a descarregar.")
                    return render(request, "course_capacity/capacity_course_upload_form.html", {"form": form})


                # Skip header row and iterate through each row in the Excel sheet
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    # Map row values to variables
                    id_course_year = row[0]
                    academic_year = row[1]
                    degree_name = row[2]
                    course_code = row[3]
                    course_name = row[4]
                    semester = row[5]
                    points_a = row[6]
                    points_b = row[7]
                    points_c = row[8]
                    points_d = row[9]
                    points_e = row[10]
                    points_f = row[11]
                    total_points = row[12]
                    language_name = row[13]
                    comment = row[14]

                    # Ensure all required fields are provided
                    if not all([academic_year, degree_name, course_code, course_name, semester,language_name]):
                        messages.warning(
                            request, f"Fila {row_num} no s'ha processat correctament: informació incompleta."
                        )
                        error_occurred = True
                        continue

                    if redirect_year is None and academic_year:
                        try:
                            year = Year.objects.get(Year=academic_year)
                            redirect_year = year.idYear 
                        except Year.DoesNotExist:
                            messages.warning(request, f"Fila {row_num} no s'ha processat: curs acadèmic '{academic_year}' no trobat.")
                            error_occurred = True
                            continue

                    if not isinstance(id_course_year, (int, float)) or not str(id_course_year).isdigit():
                        messages.warning(
                            request, f"Fila {row_num} no s'ha processat: l'ID de l'encàrrec docent ha de ser un número."
                        )
                        error_occurred = True
                        continue

                    if not isinstance(course_code, (int, float)) or not str(course_code).isdigit():
                        messages.warning(
                            request, f"Fila {row_num} no s'ha processat: el codi de l'assignatura ha de ser un número."
                        )
                        error_occurred = True
                        continue

                    points = [points_a, points_b, points_c, points_d, points_e, points_f]
                    valid_points = []
                    for idx, point in enumerate(points, start=1):
                        if point is not None:
                            try:
                                # Try converting the point to a float (valid numeric value)
                                valid_points.append(float(point))
                            except ValueError:
                                messages.warning(request, f"Fila {row_num}: El valor en la columna 'Punts' (fila {idx}) no és un número vàlid.")
                                error_occurred = True
                                continue

                    if len(valid_points) == 0:
                        messages.warning(request, f"Fila {row_num}: Cal introduir almenys un valor en les columnes 'Punts A', 'Punts B', 'Punts C', 'Punts D', 'Punts E' o 'Punts F'.")
                        error_occurred = True
                        continue
                    else:
                        total_calculated = sum(valid_points)

                        if total_points is not None and total_calculated != total_points:
                            messages.warning(request, f"Fila {row_num}: La suma dels punts no coincideix amb el valor a la columna 'Suma'. Calculat: {total_calculated}, Esperat: {total_points}.")
                            error_occurred = True
                            continue

                    # Fetch related objects
                    try:
                        year = Year.objects.get(Year=academic_year)
                    except Year.DoesNotExist:
                        messages.warning(request, f"Fila {row_num} no s'ha processat: curs acadèmic '{academic_year}' no trobat.")
                        error_occurred = True
                        continue

                    degree = Degree.objects.filter(NameDegree=degree_name).first()
                    if not degree:
                        messages.warning(
                            request,
                            f"Fila {row_num} no s'ha processat: la titulació '{degree_name}' no existeix."
                        )
                        error_occurred = True
                        continue
                    try:
                        course = Course.objects.get(CodeCourse=course_code)  # Fetch Course using course_code
                        # Check if the course name matches
                        if course.NameCourse != course_name:
                            messages.warning(request, f"Fila {row_num}: El nom de l'assignatura '{course_name}' no coincideix amb el codi '{course_code}'.")
                            error_occurred = True
                            continue
                    except Course.DoesNotExist:
                        messages.warning(request, f"Fila {row_num}: Codi d'assignatura '{course_code}' no trobat.")
                        error_occurred = True
                        continue

                    try:
                        language = Language.objects.get(Language=language_name) 
                    except Language.DoesNotExist:
                        messages.warning(request, f"Fila {row_num} no s'ha processat: idioma '{language}' no existeix.")
                        error_occurred = True
                        continue

                    # Handle semester choices
                    if semester not in ['Q1', 'Q2']:
                        messages.warning(request, f"Fila {row_num} no s'ha processat: semestre '{semester}' invàlid.")
                        error_occurred = True
                        continue
                    
                    # Create or update CourseYear
                    existing_course = CourseYear.objects.filter(idCourseYear=id_course_year).first()
                   
                    course=Course.objects.get(CodeCourse=course_code)
                    year=Year.objects.get(Year=academic_year)
                    language=Language.objects.get(Language=language_name)

                    if existing_course:
                        # Ensure no duplicate codes or names within the same degree
                        duplicate_course = CourseYear.objects.filter(
                            Course=course,
                            Year=year,
                            Semester=semester
                        ).exclude(pk=existing_course.pk)  
                        if duplicate_course.exists():
                            messages.warning(
                                request,
                                f"Fila {row_num} no s'ha processat: la combinació de l'assignatura, curs acadèmic i semestre ja existeix."
                            )
                            error_occurred = True
                            continue

                        # Update the existing course year
                        existing_course.Course = course
                        existing_course.Year=year
                        existing_course.Semester = semester
                        existing_course.PointsA=points_a
                        existing_course.PointsB=points_b
                        existing_course.PointsC=points_c
                        existing_course.PointsD=points_d
                        existing_course.PointsE=points_e
                        existing_course.PointsF=points_f 
                        existing_course.Language=language 
                        existing_course.Comment=comment
                        existing_course.save()

                    else:
                        # Ensure no duplicates for new courses
                        if CourseYear.objects.filter(Course=course,Year=year,Semester=semester).exists():
                            messages.warning(
                                request,
                                f"Fila {row_num} no s'ha processat: la combinació de l'assignatura, curs acadèmic i semestre ja existeix."
                            )
                            error_occurred = True
                            continue

                        # Create a new course
                        CourseYear.objects.create(
                            Course=course,
                            Year=year,
                            Semester=semester,
                            PointsA=points_a,
                            PointsB=points_b,
                            PointsC=points_c,
                            PointsD=points_d,
                            PointsE=points_e,
                            PointsF=points_f,
                            Language=language,
                            Comment=comment,
                        )
                    

                if not error_occurred:
                    messages.success(request, "Tots els encàrrecs docents s'han processat correctament.")
                    if redirect_year:
                            return redirect(f"{reverse('courseyear_list')}?year={redirect_year}")
                    else:
                        return redirect("courseyear_list") 

            except Exception as e:
                messages.error(request, f"Error en processar el fitxer: {e}")
                error_occurred = True

        return render(request, "course_capacity/capacity_course_upload_form.html",{"form": form})
    else:
        form = UploadForm() 

    return render(request, "course_capacity/capacity_course_upload_form.html",{"form": form})