import openpyxl
from openpyxl.styles import Alignment
from django.http import HttpResponse
from datetime import datetime
from .models import Degree,School,Course,Field
from .forms import UploadForm
from django.shortcuts import render, redirect,get_object_or_404
from django.contrib import messages

## DEGREE EXCEL

def generate_degree_excel(request):
    try:
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Titulacions"

        headers = [
            ("ID de la titulació", 15),
            ("Titulació", 30),
            ("Codi", 15),
            ("Escola", 30),
            ("Està Activa", 15),
        ]

        # headers to the first row
        for col_num, (header, width) in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        # Write the data for each Degree
        degrees = Degree.objects.all()
        for row_num, degree in enumerate(degrees, start=2):
            sheet.cell(row=row_num, column=1, value=degree.idDegree)
            sheet.cell(row=row_num, column=2, value=degree.NameDegree)
            sheet.cell(row=row_num, column=3, value=degree.CodeDegree)
            sheet.cell(row=row_num, column=4, value=degree.School.NameSchool if degree.School else "")
            sheet.cell(row=row_num, column=5, value="Si" if degree.isActive else "No")

        # Generate a timestamped file name
        current_datetime = datetime.now().strftime("%d%m%Y_%H%M%S")
        filename = f"Titulacions_{current_datetime}.xlsx"

        # Set up the HTTP response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        workbook.save(response)

        return response
    except Exception as e:
        messages.error(request, f"Error al generar el fitxer Excel: {str(e)}")
        return redirect("course_list")

def upload_degree_excel(request):
    if request.method == "POST":
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["file"]
            error_occurred = False
            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                expected_columns =["ID de la titulació","Titulació","Codi","Escola","Està Activa"]
                header_row = [cell.value for cell in sheet[1]]

                # Check columns
                if not all(col in header_row for col in expected_columns):
                    messages.error(request, "El fitxer no té les columnes esperades. Comprova que el fitxer segueixi les dades com el excel a descarregar.")
                    return render(request, "degree_upload_form.html", {"form": form})


                #skip header
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    id_degree = row[0]
                    name_degree = row[1]
                    code_degree = row[2]
                    school_name = row[3]
                    is_active = row[4]

                    if not all([id_degree, name_degree, code_degree, school_name]):
                        messages.warning(
                            request, f"Fila {row_num} no s'ha processat correctament: informació incompleta."
                        )
                        error_occurred=True
                        continue

                    if not isinstance(id_degree, (int, float)) or not str(id_degree).isdigit():
                        messages.warning(
                            request, f"Fila {row_num} no s'ha processat: l'ID de titulació ha de ser un número."
                        )
                        error_occurred = True
                        continue

                    if not isinstance(code_degree, (int, float)) or not str(code_degree).isdigit():
                        messages.warning(
                            request, f"Fila {row_num} no s'ha processat: el codi de titulació ha de ser un número."
                        )
                        error_occurred = True
                        continue
                    
                    if isinstance(is_active, str):
                        if is_active == 'Si':
                            is_active = True
                        elif is_active == 'No':
                            is_active = False
                        else:
                            messages.warning(request, f"Fila {row_num} no s'ha processat: l'estat d'activitat ha de ser 'Si' o 'No'.")
                            error_occurred = True
                            continue


                    try:
                        school = School.objects.get(NameSchool=school_name)
                    except School.DoesNotExist:
                        messages.warning(
                            request, f"Fila {row_num} no s'ha processat: l'escola '{school_name}' no existeix."
                        )
                        error_occurred=True
                        continue

                    #If exists
                    existing_degree = Degree.objects.filter(idDegree=id_degree).first()

                    if existing_degree:
                        if Degree.objects.filter(NameDegree=name_degree, School=school).exclude(idDegree=existing_degree.idDegree).exists():
                            messages.warning(
                                request, f"Fila {row_num} no s'ha processat: la combinació de titulació i escola ja existeix."
                            )
                            error_occurred = True
                            continue
                        if Degree.objects.filter(CodeDegree=code_degree).exclude(CodeDegree=existing_degree.CodeDegree).exists():
                            messages.warning(
                                request, f"Fila {row_num} no s'ha processat: el codi de la titulació ja existeix."
                            )
                            error_occurred = True
                            continue
                        # Update the existing degree
                        existing_degree.NameDegree = name_degree
                        existing_degree.CodeDegree = code_degree
                        existing_degree.School = school
                        existing_degree.isActive = is_active
                        existing_degree.save()

                    #Does not exist - create
                    else:
                        if Degree.objects.filter(CodeDegree=code_degree).exists():
                            messages.warning(
                                request, f"Fila {row_num} no s'ha processat: el codi de la titulació ja existeix."
                            )
                            error_occurred = True
                            continue

                        if Degree.objects.filter(NameDegree=name_degree, School=school).exists():
                            messages.warning(
                                request, f"Fila {row_num} no s'ha processat: la combinació de titulació i escola ja existeix."
                            )
                            error_occurred = True
                            continue

                        # Create a new degree
                        degree = Degree.objects.create(
                            idDegree=id_degree,
                            NameDegree=name_degree,
                            CodeDegree=code_degree,
                            School=school,
                            isActive=is_active,
                        )
                
                if not error_occurred:
                    messages.success(request, "Les titulaciosn s'han actualitzat correctament.")
                    return redirect('degree_list')

            except Exception as e:
                messages.error(request, f"Error en processar el fitxer: {e}")
                error_occurred=True

            return render(request, "degree_upload_form.html", {"form": form})

    else:
        form = UploadForm()

    return render(request, "degree_upload_form.html", {"form": form})

## COURSES EXCEL

def generate_course_excel(request):
    try:
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Assignatures"

        headers = [
            ("ID d'assignatura", 15),
            ("Titulació", 30),
            ("Nom de l'assignatura", 30),
            ("Codi", 15),
            ("ECTS", 15),
            ("Camp de coneixement", 30),
            ("Està Activa", 15),
        ]

        # Set headers in the first row
        for col_num, (header, width) in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        # Write course data
        courses = Course.objects.select_related("Degree", "Field").order_by(
            "Degree__NameDegree", "NameCourse"
        )
        for row_num, course in enumerate(courses, start=2):
            sheet.cell(row=row_num, column=1, value=course.idCourse)
            sheet.cell( row=row_num,column=2,value=course.Degree.NameDegree if course.Degree else "")
            sheet.cell(row=row_num, column=3, value=course.NameCourse)
            sheet.cell(row=row_num, column=4, value=course.CodeCourse)
            sheet.cell(row=row_num, column=5, value=course.ECTS)
            sheet.cell(row=row_num,column=6, value=course.Field.NameField if course.Field else "")
            sheet.cell(row=row_num, column=7, value="Si" if course.isActive else "No")

        # Generate a timestamped file name
        current_datetime = datetime.now().strftime("%d%m%Y_%H%M%S")
        filename = f"Assignatures_{current_datetime}.xlsx"

        # Return Excel file as HTTP response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        workbook.save(response)

        return response
    except Exception as e:
        messages.error(request, f"Error al generar el fitxer  Excel: {str(e)}")
        return redirect("course_list")


def upload_course_excel(request):
    if request.method == "POST":
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["file"]
            error_occurred = False
            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                expected_columns =["ID d'assignatura","Titulació","Nom de l'assignatura","Codi","ECTS","Camp de coneixement","Està Activa",]
                header_row = [cell.value for cell in sheet[1]]

                # Check columns
                if not all(col in header_row for col in expected_columns):
                    messages.error(request, "El fitxer no té les columnes esperades. Comprova que el fitxer segueixi les dades com el excel a descarregar.")
                    return render(request, "course_upload_form.html", {"form": form})

                
                # Process rows, skipping the header
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    id_course = row[0]
                    degree_name = row[1]
                    name_course = row[2]
                    code_course = row[3]
                    ects = row[4]
                    field_name = row[5]
                    is_active = row[6]
                    
                  # Check for missing required fields
                    if not all([id_course, degree_name, name_course, code_course, ects,field_name,is_active]):
                        messages.warning(
                            request,
                            f"Fila {row_num} no s'ha processat correctament: informació incompleta."
                        )
                        error_occurred = True
                        continue

                    # Validate ID and Code 
                    if not isinstance(id_course, (int, float)) or not str(id_course).isdigit():
                        messages.warning(
                            request, f"Fila {row_num} no s'ha processat: l'ID de l'assignatura ha de ser un número."
                        )
                        error_occurred = True
                        continue

                    if not isinstance(code_course, (int, float)) or not str(code_course).isdigit():
                        messages.warning(
                            request, f"Fila {row_num} no s'ha processat: el codi de l'assignatura ha de ser un número."
                        )
                        error_occurred = True
                        continue

                    if not isinstance(ects, (int, float)) or not str(ects).isdigit():
                        messages.warning(
                            request, f"Fila {row_num} no s'ha processat: els etcs han de ser un número."
                        )
                        error_occurred = True
                        continue


                    if is_active not in ["Si", "No"]:
                        messages.warning(
                            request, f"Fila {row_num} no s'ha processat: l'estat d'activitat ha de ser 'Si' o 'No'."
                        )
                        error_occurred = True
                        continue

                    # Get related Degree and Field
                    degree = Degree.objects.filter(NameDegree=degree_name).first()
                    if not degree:
                        messages.warning(
                            request,
                            f"Fila {row_num} no s'ha processat: la titulació '{degree_name}' no existeix."
                        )
                        error_occurred = True
                        continue

                    field = Field.objects.filter(NameField=field_name).first()
                    if not field:
                        messages.warning(
                            request,
                            f"Fila {row_num} no s'ha processat: el camp de coneixement '{field_name}' no existeix."
                        )
                        error_occurred = True
                        continue

                    # Check if the course already exists
                    existing_course = Course.objects.filter(idCourse=id_course).first()

                    if existing_course:
                        # Ensure no duplicate codes or names within the same degree
                        if Course.objects.filter(
                            NameCourse=name_course, Degree=degree
                        ).exclude(idCourse=existing_course.idCourse).exists():
                            messages.warning(
                                request,
                                f"Fila {row_num} no s'ha processat: la combinació de l'assignatura i titulació ja existeix."
                            )
                            error_occurred = True
                            continue

                        if Course.objects.filter(CodeCourse=code_course).exclude(idCourse=existing_course.idCourse).exists():
                            messages.warning(
                                request,
                                f"Fila {row_num} no s'ha processat: el codi de l'assignatura ja existeix."
                            )
                            error_occurred = True
                            continue

                        # Update the existing course
                        existing_course.NameCourse = name_course
                        existing_course.CodeCourse = code_course
                        existing_course.ECTS = ects
                        existing_course.Degree = degree
                        existing_course.Field = field
                        existing_course.isActive = str(is_active).strip().lower() == "si"
                        existing_course.save()
                    else:
                        # Ensure no duplicates for new courses
                        if Course.objects.filter(CodeCourse=code_course).exists():
                            messages.warning(
                                request,
                                f"Fila {row_num} no s'ha processat: el codi de l'assignatura ja existeix."
                            )
                            error_occurred = True
                            continue

                        if Course.objects.filter(NameCourse=name_course, Degree=degree).exists():
                            messages.warning(
                                request,
                                f"Fila {row_num} no s'ha processat: la combinació de l'assignatura i titulació ja existeix."
                            )
                            error_occurred = True
                            continue

                        # Create a new course
                        Course.objects.create(
                            idCourse=id_course,
                            NameCourse=name_course,
                            CodeCourse=code_course,
                            ECTS=ects,
                            Degree=degree,
                            Field=field,
                            isActive=str(is_active).strip().lower() == "si",
                        )

                if not error_occurred:
                    messages.success(request, "Les assignatures s'han actualitzat correctament.")
                    return redirect("course_list")

            except Exception as e:
                messages.error(request, f"Error en processar el fitxer: {e}")
                error_occurred = True

            return render(request, "course_upload_form.html", {"form": form})

    else:
        form = UploadForm()

    return render(request, "course_upload_form.html", {"form": form})