# services.py
from django.contrib.auth.hashers import make_password
from django.contrib import messages
import pandas as pd
from UsersApp.models import CustomUser, Professor,ProfessorField,ProfessorLanguage,TypeProfessor,Chief
from AcademicInfoApp.models import Field,Language,Section
import openpyxl
from openpyxl.styles import Alignment
from django.http import HttpResponse
from datetime import datetime
from .forms import UploadForm
from django.shortcuts import render, redirect,get_object_or_404
from django.contrib import messages

def generate_professor_excel(request):
    try:
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Professorat"

        headers = [
            ("ID del professor", 15),
            ("Username",20),
            ("Nom", 20),
            ("Cognoms", 30),
            ("Correu",30),
            ("Rol",15),
            ("Secció",18),
            ("Contracte actual",25),
            ("Camps de coneixement",30),
            ("Idiomes",30),
            ("Està Actiu", 9),
            ("Comentari cap de secció", 50),
            ("Descripció direcció", 50),
        ]

        # Add headers to the first row
        for col_num, (header, width) in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        # Write the data for each professor
        professors = Professor.objects.all().order_by("family_name")

        for row_num, professor in enumerate(professors, start=2):
            # Get related fields, languages, and contract
            fields = ProfessorField.objects.filter(Professor=professor)
            languages = ProfessorLanguage.objects.filter(Professor=professor)
            contract = professor.current_contract

            # Join multiple fields/languages with commas, handle nulls
            fields_str = ", ".join(field.Field.NameField for field in fields) if fields.exists() else ""
            languages_str = ", ".join(lang.Language.Language for lang in languages) if languages.exists() else ""
            contract_str = contract.NameContract if contract else ""

            max_length = 350  
            truncated_comment = (professor.comment[:max_length] + "...") if professor.comment and len(professor.comment) > max_length else professor.comment
            truncated_description = (professor.description[:max_length] + "...") if professor.description and len(professor.description) > max_length else professor.description

            section_section_chief = ""
            if professor.user.role == "section_chief":
                chief = Chief.objects.get(professor=professor)
                section_section_chief = chief.section.NameSection if chief.section else ""    

            sheet.cell(row=row_num, column=1, value=professor.idProfessor).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=2, value=professor.user.username if professor.user else "").alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=3, value=professor.name).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=4, value=professor.family_name).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=5, value=professor.email).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=6, value="professor" if professor.user.role == "professor" else "cap de secció" if professor.user.role == "section_chief" else "").alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=7, value=section_section_chief).alignment = Alignment(horizontal="left")  
            sheet.cell(row=row_num, column=8, value=contract_str).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=9, value=fields_str).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=10, value=languages_str).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=11, value="Si" if professor.isActive else "No").alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=12, value=truncated_comment or "").alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=13, value=truncated_description or "").alignment = Alignment(horizontal="left")


        # Generate a timestamped file name
        current_datetime = datetime.now().strftime("%d%m%Y_%H%M%S")
        filename = f"Professorat_{current_datetime}.xlsx"

        # Set up the HTTP response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        workbook.save(response)

        return response
    except Exception as e:
        messages.error(request, f"Error al generar el fitxer Excel: {str(e)}")
        return redirect("usersapp:professor_list")

def upload_professor_excel(request):
    if request.method == "POST":
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["file"]
            error_occurred = False

            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                # Expected columns in the header row
                expected_columns = [
                    "ID del professor",
                    "Username",
                    "Nom",
                    "Cognoms",
                    "Correu",
                    "Rol",
                    "Secció",
                    "Contracte actual",
                    "Camps de coneixement",
                    "Idiomes",
                    "Està Actiu",
                    "Comentari cap de secció",
                    "Descripció direcció",
                ]
                header_row = [cell.value for cell in sheet[1]]

                # Validate header row
                if header_row != expected_columns:
                    messages.error(request, "El fitxer no té les columnes esperades. Comprova que el fitxer segueixi les dades com el excel a descarregar.")
                    return render(request, "users/professor/professor_upload_form.html", {"form": form})

                # Process each row
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    id_professor = row[0]
                    username = row[1]
                    name = row[2]
                    family_name = row[3]
                    email = row[4]
                    role=row[5]
                    section=row[6]
                    contract_name = row[7]
                    fields = row[8]
                    languages = row[9]
                    is_active = row[10]
                    comment = row[11]
                    description = row[12]

                    create_sectorchief=False

                    if not all([id_professor, username, name, family_name, email,role,is_active]):
                        messages.warning(request, f"Fila {row_num}: no s'ha processat correctament: informació incompleta.")
                        error_occurred = True
                        continue

                    if not isinstance(id_professor, (int, float)) or not str(id_professor).isdigit():
                        messages.warning( request, f"Fila {row_num} no s'ha processat: l'ID del professor ha de ser un número." )
                        error_occurred = True
                        continue

                    if role not in ["professor", "cap de secció"]:
                        messages.warning( request, f"Fila {row_num} no s'ha processat: el rol ha de ser 'professor' o 'cap de secció'." )
                        error_occurred = True
                        continue

                    actual_role = "section_chief" if role == "cap de secció" else "professor"

                    if role=="cap de secció":                    
                        exist_section = Section.objects.filter(NameSection=section).first()
                        if exist_section is None:
                            messages.warning(request, f"Fila {row_num}: El professor té el rol assignat de '{role}'en la secció '{section}' que no existeix.")
                            error_occurred = True
                            continue

                        # Check if the section already has a chief assigned
                        current_chief = Chief.objects.filter(section=exist_section).first()
                        professor = Professor.objects.filter(idProfessor=id_professor).first()
                        if current_chief and current_chief.professor != professor:
                            messages.warning(request, f"Fila {row_num}: La secció '{section}' ja té un cap de secció assignat.")
                            error_occurred = True
                            continue

                        create_sectorchief = True
                    else:
                        create_sectorchief = False
                    

                    if is_active not in ["Si", "No"]:
                        messages.warning( request, f"Fila {row_num} no s'ha processat: l'estat d'activitat ha de ser 'Si' o 'No'." )
                        error_occurred = True
                        continue

                    is_active = is_active.lower() == "si"

                    if contract_name:
                        if not TypeProfessor.objects.filter(NameContract=contract_name).exists():
                            messages.warning(request, f"Fila {row_num}: El tipus de contracte '{contract_name}' no existeix.")
                            contract = None
                        else:
                            contract = TypeProfessor.objects.filter(NameContract=contract_name).first()
                    else:
                        contract = None

                    missing_fields = []
                    valid_fields = []
                    for field_name in (fields or '').split(','):
                        field_name = field_name.strip()
                        if field_name:
                            field = Field.objects.filter(NameField=field_name).first()
                            if not field:
                                missing_fields.append(field_name)
                            else:
                                valid_fields.append(field)

                    missing_languages = []
                    valid_languages = []
                    for language_name in (languages or '').split(','):
                        language_name = language_name.strip()
                        if language_name:
                            language = Language.objects.filter(Language=language_name).first()
                            if not language:
                                missing_languages.append(language_name)
                            else:
                                valid_languages.append(language)

                    if missing_fields or missing_languages:
                        messages.warning(request, f"Fila {row_num}: Alguns camps o idiomes no existeixen: {', '.join(missing_fields + missing_languages)}.")


                    # Check for existing professor
                    professor = Professor.objects.filter(idProfessor=id_professor).first()
                    if professor:

                        if CustomUser.objects.exclude(pk=professor.user.id).filter(username=username).exists():
                            messages.warning(request, f"Fila {row_num}: El nom d'usuari '{username}' ja existeix.")
                            error_occurred = True
                            continue

                        if CustomUser.objects.exclude(pk=professor.user.id).filter(email=email).exists():
                            messages.warning(request, f"Fila {row_num}: El correu '{email}' ja està en ús.")
                            error_occurred = True
                            continue

                        # Update existing professor
                        user=professor.user
                        user.username = username
                        user.email = email
                        user.first_name= name
                        user.last_name = family_name
                        user.role=actual_role

                        professor.name = name
                        professor.family_name = family_name
                        professor.email = email
                        professor.current_contract = contract
                        professor.comment = comment
                        professor.description = description
                        professor.isActive = is_active
                        
                        user.save()
                        professor.save()                       

                    else:
                        # Create new professor
                        if Professor.objects.filter(idProfessor=id_professor).exists():
                            messages.warning(request, f"Fila {row_num}: L'id del Professor '{id_professor}' ja existeix.")
                            error_occurred = True
                            continue

                        if CustomUser.objects.filter(username=username).exists():
                            messages.warning(request, f"Fila {row_num}: El nom d'usuari '{username}' ja està en ús.")
                            error_occurred = True
                            continue

                        if CustomUser.objects.filter(email=email).exists():
                            messages.warning(request, f"Fila {row_num}: El correu '{email}' ja està en ús.")
                            error_occurred = True
                            continue

                        # Create new professor and user
                        try:
                            user = CustomUser.objects.create_user(username=username, first_name=name, last_name=family_name, email=email, password="default_password",is_active=is_active,role=actual_role)
                            professor = Professor.objects.create(
                                idProfessor=id_professor,
                                user=user,
                                name=name,
                                family_name=family_name,
                                email=email,
                                current_contract=contract,
                                comment=comment,
                                description=description,
                                isActive = is_active,
                            )


                        except Exception as e:
                            messages.warning(request, f"Fila {row_num}: Error al crear el professor. {e}")
                            error_occurred = True
                            continue

                    # Update related fields and languages
                    ProfessorField.objects.filter(Professor=professor).delete()
                    for field in valid_fields:
                        ProfessorField.objects.create(Professor=professor, Field=field)

                    ProfessorLanguage.objects.filter(Professor=professor).delete()
                    for language in valid_languages:
                        ProfessorLanguage.objects.create(Professor=professor, Language=language)

                    if create_sectorchief:
                        try:
                            existing_chief = Chief.objects.filter(section=exist_section).first()
                            if existing_chief:
                                if existing_chief.professor != professor:
                                    messages.warning( request, f"Fila {row_num}: La secció '{section}' ja té un cap de secció assignat.")
                                    error_occurred = True
                            else:
                                Chief.objects.create(professor=professor, section=exist_section)
                        except Exception as e:
                            messages.warning(request, f"Fila {row_num}: Error al crear el cap de secció per al professor '{professor.name}': {str(e)}")
                            error_occurred = True
                
                if not error_occurred:
                    messages.success(request, "Els professors s'han actualitzat correctament.")
                    return redirect("usersapp:professor_list")

            except Exception as e:
                messages.error(request, f"Error en processar el fitxer: {e}")
                error_occurred = True

            return render(request, "users/professor/professor_upload_form.html", {"form": form})

    else:
        form = UploadForm()

    return render(request, "users/professor/professor_upload_form.html", {"form": form})
